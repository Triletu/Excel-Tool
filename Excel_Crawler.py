from openpyxl.styles import colors
from openpyxl.styles import Font, Color
import openpyxl
from tkinter import *
from tkinter.ttk import *
import tkinter.filedialog as fdialog


class Excel_Crawler:

	def method_name(self):
	    return self.label_key_column

	def __init__(self, master):
		frame = Frame(master)
		frame.pack()

		self.combo_1_act_val = StringVar()
		self.combo_2_act_val = StringVar()
		self.label_input = Label(frame, text='Select Input File  ')
		self.label_input.grid(row=0, sticky = E, pady=10, padx=10)
		self.label_output = Label(frame, text='Select Output File  ')
		self.label_output.grid(row=1, sticky = E, padx=10)
		self.button_input = Button(frame, text='Input', command = self.assign_input, width = 10)
		self.button_input.grid(row=0, column=1, sticky=W, pady=10)
		self.button_output = Button(frame, text='Output', command = self.assign_output, width = 10)
		self.button_output.grid(row=1, column=1, sticky=W,)
		self.label_file_1_name = Label(frame, text='  File name:  ')
		self.label_file_1_name.grid(row=0, column=2, sticky = E, pady=10, padx=10)
		self.label_file_2_name = Label(frame, text='  File name:  ')
		self.label_file_2_name.grid(row=1, column=2, sticky = E,padx=10)
		self.button_start = Button(frame, text='Start!', command = self.go, width = 10)
		self.button_start.grid(row=2, column=7, pady=20)
		self.label_combo_1 = Label(frame, text='Sheet:')
		self.label_combo_1.grid(row=0, column=3, sticky = E, pady=10, padx=10)
		self.label_combo_2 = Label(frame, text='Sheet:')
		self.label_combo_2.grid(row=1, column=3, sticky = E, padx=10)
		self.combo_file_1 = Combobox(frame, textvariable=self.combo_1_act_val, values = 'Select_File_First')
		self.combo_file_1.grid(row=0, column=4, padx = 10)
		self.combo_file_1.state(['readonly'])
		self.combo_file_2 = Combobox(frame, textvariable=self.combo_2_act_val, values = 'Select_File_First' )
		self.combo_file_2.grid(row=1, column=4, padx = 10)
		self.combo_file_2.state(['readonly'])
		self.label_key_column1 = Label(frame, text='Key Column:')
		self.label_key_column1.grid(row=0, column=5, padx = 10)
		self.entry_key_column1 = Entry(frame, width=5)
		self.entry_key_column1.grid(row=0, column=6, padx=10)
		self.label_data_column1 = Label(frame, text='Data Column:')
		self.label_data_column1.grid(row=0, column=7, padx = 10)
		self.entry_data_column1 = Entry(frame, width=5)
		self.entry_data_column1.grid(row=0, column=8, padx=10)
		self.label_key_column2 = Label(frame, text='Key Column:')
		self.label_key_column2.grid(row=1, column=5, padx = 10)
		self.entry_key_column2 = Entry(frame, width=5)
		self.entry_key_column2.grid(row=1, column=6, padx=10)
		self.label_data_column2 = Label(frame, text='Data Column:')
		self.label_data_column2.grid(row=1, column=7, padx = 10)
		self.entry_data_column2 = Entry(frame, width=5)
		self.entry_data_column2.grid(row=1, column=8, padx=10)
		self.entry_key_column1.insert(END,'C')
		self.entry_data_column1.insert(END,'D')
		self.entry_key_column2.insert(END,'E')
		self.entry_data_column2.insert(END,'J')
		
	def assign_input(self):
		file = fdialog.askopenfilename()  
		self.wb_inp = openpyxl.load_workbook(file)
		self.label_file_1_name['text']= str(file)
		sheet_names = self.wb_inp.get_sheet_names()
		self.combo_file_1['values'] = sheet_names
		
	def assign_output(self):
		file = fdialog.askopenfilename()
		self.wb_out = openpyxl.load_workbook(file)	
		self.label_file_2_name['text']= str(file)
		sheet_names = self.wb_out.get_sheet_names()
		self.combo_file_2['values'] = sheet_names

	def go(self):
		self.sheet_in = self.wb_inp.get_sheet_by_name(self.combo_file_1.get())
		self.sheet_out = self.wb_out.get_sheet_by_name(self.combo_file_2.get())
		input_values = {'':''}
		ft = Font(color=colors.RED)
		for i in range(1, self.sheet_in.max_row + 1):
			input_values[self.sheet_in[str(self.entry_key_column1.get())+str(i)].value] = self.sheet_in[str(self.entry_data_column1.get())+str(i)].value
			
		for x in range(1, self.sheet_out.max_row + 1):
			if self.sheet_out[str(self.entry_key_column2.get())+str(x)].value in input_values:
				description = input_values[self.sheet_out[str(self.entry_key_column2.get())+str(x)].value]
				self.sheet_out[str(self.entry_data_column2.get())+str(x)].value = description
				self.wb_out.save('completed.xlsx')
			else:
				self.sheet_out[self.entry_key_column2.get()+str(x)].font = ft

Tk_window = Tk()

Tk_window.title("Excel Crawler")
start = Excel_Crawler(Tk_window)
Tk_window.mainloop()